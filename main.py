import os
import urllib.parse
import datetime
import time
import selenium.common
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import openpyxl


def login(username, password):
    # 输入用户名/密码登录
    driver.get("http://login.sina.com.cn/")
    elem_user = driver.find_element_by_name("username")
    elem_user.send_keys(username)  # 用户名
    elem_pwd = driver.find_element_by_name("password")
    elem_pwd.send_keys(password)  # 密码

    # elem_sub = driver.find_element_by_xpath("//input[@class='smb_btn']")
    # elem_sub.click()              #点击登陆
    # 如果登陆按钮采用动态加载 则采用输入回车键登陆微博
    elem_pwd.send_keys(Keys.RETURN)
    time.sleep(2)


def is_element_exists(driver, css):
    try:
        driver.find_element_by_css_selector(css)
        return True
    except selenium.common.exceptions.NoSuchElementException:
        return False


string = input("检索关键词:")
stringIn = urllib.parse.quote(urllib.parse.quote(string))

start = 1
row = 0
if os.path.isfile(string+".xlsx"):
    file = openpyxl.load_workbook(string+".xlsx")
    ws = file.get_sheet_by_name(file.get_sheet_names()[0])
    rows = []
    for row in ws.rows:
        rows.append(row)
    row = len(rows)
    start = row + 1
else:
    file = openpyxl.Workbook()
    ws = file.get_active_sheet()

pretime = datetime.datetime(2017, 5, 26)  # 起始时间
afttime = datetime.datetime(2017, 12, 31)  # 结束时间
today = datetime.datetime.now()
if row != 0:
    a = int(ws.cell(row=row, column=1).value.split(".")[0])
    b = int(ws.cell(row=row, column=1).value.split(".")[1])
    c = int(ws.cell(row=row, column=1).value.split(".")[2])
    if pretime < datetime.datetime(a, b, c):
        pretime = datetime.datetime(a, b, c)

driver = webdriver.Chrome()
driver.implicitly_wait(30)

login("your email", "your password")  # 在这里修改密码

while pretime <= afttime:
    url = "https://s.weibo.com/weibo/" + stringIn + "&typeall=1&suball=1&timescope=custom:" + \
          pretime.strftime("%Y-%m-%d") + ":" + pretime.strftime("%Y-%m-%d") + "&page="
    for i in range(1, 51):
        driver.get(url+str(i))
        WebDriverWait(driver, 20, 0.5).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "help_link")))
        # now_handle = driver.window_handles[0]
        if "你的行为有些异常，请输入验证码" in driver.page_source:  # 验证码检测
            print("请输入验证码!!!")
            WebDriverWait(driver, 60000, 0.5).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "comment_txt")))  # 等待输入验证码
        WebDriverWait(driver, 20, 0.5).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "help_link")))
        WebDriverWait(driver, 20, 0.5).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "search_feed")))
        test = driver.find_element_by_css_selector(".search_feed")
        if is_element_exists(test, ".search_noresult"):  # 页数超出范围或没有内容
            print("page " + str(i) + " no index")
            break
        content = driver.find_element_by_css_selector(".feed_lists.W_texta")
        sources = content.find_elements_by_css_selector(".WB_cardwrap.S_bg2.clearfix")
        for source in sources:
            time = source.find_elements_by_css_selector(".feed_from.W_textb")[-1].text.split("来自")[0]
            print(time)
            if "前" in time:  # 前xx分钟的格式
                print(str(today.month) + "月" + str(today.day) + "日")
                now_year = today.year
                now_month = today.month
                now_day = today.day
                pre_minute = int(time.split("分钟")[0])
                now_time = today - datetime.timedelta(minutes=pre_minute)
                print(now_time.strftime("%H:%M"))
            else:
                if "今天" in time:
                    print(str(today.month)+"月"+str(today.day)+"日")
                    now_year = today.year
                    now_month = today.month
                    now_day = today.day
                    time = time.split("今天")[1].strip().split(" ")[0]
                else:
                    if "月" in time and "日" in time:
                        print(time.split(" ")[0])
                        now_year = today.year
                        now_month = int(time.split(" ")[0].split("月")[0])
                        now_day = int(time.split(" ")[0].split("月")[1].split("日")[0])
                        time = time.split("日")[1].strip().split(" ")[0]
                        print(time)
                    else:
                        now_year = int(time.split(" ")[0].split("-")[0])
                        now_month = int(time.split(" ")[0].split("-")[1])
                        now_day = int(time.split(" ")[0].split("-")[2])
                        print(time.split(" ")[0])
                        time = time.split(" ")[1]
                        print(time)
                now_time = datetime.datetime.strptime(time, '%H:%M')

            # 比较时间，若早于最后一条，则跳过
            if row != 0:
                last_year = int(ws.cell(row=row, column=1).value.split(".")[0])
                last_month = int(ws.cell(row=row, column=1).value.split(".")[1])
                last_day = int(ws.cell(row=row, column=1).value.split(".")[2])
                if datetime.datetime(now_year, now_month, now_day) < datetime.datetime(last_year, last_month, last_day):
                    break
                elif datetime.datetime(today.year, now_month, now_day) == datetime.datetime(last_year, last_month, last_day):
                    if now_time > datetime.datetime.strptime(ws.cell(row=row, column=2).value, '%H:%M'):
                        continue
                    elif now_time == datetime.datetime.strptime(ws.cell(row=row, column=2).value, '%H:%M'):
                        if source.find_element_by_css_selector(".W_texta.W_fb").text == ws.cell(row=row, column=3).value:
                            if source.find_element_by_class_name("comment_txt").text == ws.cell(row=row, column=4).value:
                                continue

            ws.cell(row=start, column=1).value = str(now_year) + "." + str(now_month) + "." + str(now_day)
            ws.cell(row=start, column=2).value = now_time.strftime("%H:%M")
            print(source.find_element_by_css_selector(".W_texta.W_fb").text)
            ws.cell(row=start, column=3).value = source.find_element_by_css_selector(".W_texta.W_fb").text
            print(source.find_element_by_class_name("comment_txt").text)
            ws.cell(row=start, column=6).value = source.find_element_by_class_name("comment_txt").text

            url_user = source.find_element_by_css_selector(".W_texta.W_fb").get_attribute('href')

            js = 'window.open("' + url_user + '");'
            driver.execute_script(js)
            handles = driver.window_handles
            for handle in handles:
                if handle != driver.current_window_handle:
                    driver.switch_to.window(handle)
            WebDriverWait(driver, 20, 0.5).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "footer_link")))

            # 获取性别
            sex_content = driver.find_element_by_css_selector(".pf_username")
            if is_element_exists(sex_content, ".icon_pf_male"):
                print("男")
                ws.cell(row=start, column=4).value = "男"
            else:
                print("女")
                ws.cell(row=start, column=4).value = "女"
            # sex = driver.find_elements_by_css_selector(".S_txt1.t_link")[0]
            # if '他' in sex.text:
            #
            # else:


            # 获取地址
            location_test = driver.find_elements_by_css_selector(".item.S_line2.clearfix")
            for lt in location_test:
                if is_element_exists(lt, ".W_ficon.ficon_cd_place.S_ficon"):
                    location = lt.find_element_by_css_selector(".item_text.W_fl")
                    print(location.text)
                    ws.cell(row=start, column=5).value = location.text
                    break

            driver.close()
            handles = driver.window_handles
            driver.switch_to.window(handles[0])

            start = start + 1

            file.save(string+".xlsx")

    pretime = pretime + datetime.timedelta(days=1)

driver.close()
